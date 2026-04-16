# -*- coding: utf-8 -*-
# Copyright (C) 2026 Majid Jazebi
# Web: Geotechzone.com
# All rights reserved.
# This plugin is licensed under GNU GPL v3.
"""
Excel Management Module for Advanced Import Photos
Handles Excel import/export operations for photo management. 
"""

import os
import tempfile
import warnings
from datetime import datetime

# Excel import/export (optional — required only for Excel export/import features)
from .dependency_manager import safe_import

_openpyxl = safe_import("openpyxl")
EXCEL_AVAILABLE = _openpyxl is not None

if EXCEL_AVAILABLE:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

from qgis.PyQt.QtWidgets import QMessageBox, QFileDialog, QWidget, QApplication
from qgis.core import Qgis, QgsProject, QgsMessageLog, QgsCoordinateTransform, QgsCoordinateReferenceSystem



class ExcelManager:
    """Manages Excel file operations for photo data import/export."""

    def __init__(self, iface, main_plugin=None):
        self.iface = iface
        self.main_plugin = main_plugin  # Reference to main plugin for calling apply_all_settings

    def _get_excel_file_path(self):
        """Get the path to the Excel management file in the current QGIS project directory."""
        project = QgsProject.instance()
        project_dir = os.path.dirname(project.fileName()) if project.fileName() else None

        if not project_dir:
            # If no project is saved, use the desktop
            from qgis.PyQt.QtCore import QStandardPaths
            project_dir = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)

        return os.path.join(project_dir, "photo_management.xlsx")

    def _create_excel_file(self, file_path):
        """Create a new Excel file with the required structure."""
        if not EXCEL_AVAILABLE:
            return
        # SAME FIX AS EXPORT: Use minimal template to avoid Workbook() XML conflict
        self._create_minimal_excel_template(file_path)
        
        # Load and populate the template
        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = "Photo Management"

        # Add title row
        title_cell = ws.cell(row=1, column=1, value="Visibility")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        # Merge cells for the title (assuming 11 columns now with photo_time)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

        # Add headers (now starting from row 2)
        headers = ["Photo Name", "File Path", "Group", "X Coordinate", "Y Coordinate", "Direction", "Photo Time", "Import Date", "Last Update", "Visible", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Set column widths
        column_widths = [20, 50, 20, 15, 15, 12, 20, 20, 20, 10, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        wb.save(file_path)
        wb.close()

    def _is_excel_file_open(self, file_path):
        """Check if Excel file is currently open by trying to open it in write mode."""
        import sys
        if sys.platform != 'win32':
            return False
        try:
            with open(file_path, 'r+b') as f:
                import msvcrt
                msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
            return False
        except (IOError, OSError):
            return True

    def _create_minimal_excel_template(self, file_path):
        """Create a minimal Excel file without using Workbook() to avoid XML conflicts."""
        import zipfile
        
        # Create minimal Excel file structure manually
        # This is a valid minimal .xlsx file created as a ZIP archive
        
        # Minimal content.xml for worksheet
        sheet_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheetData/>
</worksheet>'''
        
        # Minimal workbook.xml
        workbook_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/></sheets>
</workbook>'''
        
        # Minimal _rels/.rels
        rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''
        
        # Minimal xl/_rels/workbook.xml.rels
        workbook_rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>'''
        
        # [Content_Types].xml
        content_types_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>'''
        
        # Create ZIP file
        with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('[Content_Types].xml', content_types_xml)
            zf.writestr('_rels/.rels', rels_xml)
            zf.writestr('xl/workbook.xml', workbook_xml)
            zf.writestr('xl/_rels/workbook.xml.rels', workbook_rels_xml)
            zf.writestr('xl/worksheets/sheet1.xml', sheet_xml)

    def export_plugin_state(self, photo_layer, dlg=None):
        """Export all photos from the imported photos tab to Excel with visibility status."""
        if not EXCEL_AVAILABLE:
            self.iface.messageBar().pushMessage(
                "Missing Dependency",
                "Excel export requires the openpyxl library. Install it with: pip install openpyxl",
                level=Qgis.Warning,
                duration=10
            )
            return

        if not photo_layer or not photo_layer.isValid():
            QMessageBox.warning(self.iface.mainWindow(), "Export Error", "No photo layer available to export.")
            return

        # Get save location from user
        save_path, _ = QFileDialog.getSaveFileName(
            self.iface.mainWindow(),
            "Export Photo Management File",
            "photo_management_export.xlsx",
            "Excel files (*.xlsx);;All files (*.*)"
        )

        if not save_path:
            return

        try:
            # Create Excel file with all photos from imported photos tab
            # WORKAROUND: Create minimal Excel template without using Workbook()
            # This avoids the XML library conflict causing access violations in QGIS
            import zipfile
            from io import BytesIO
            
            temp_path = os.path.join(tempfile.gettempdir(), f"qgis_template_{os.getpid()}.xlsx")
            
            # Create minimal valid Excel file structure manually
            self._create_minimal_excel_template(temp_path)
            
            # Now load the template - this is safe
            wb = load_workbook(temp_path)
            ws = wb.active
            ws.title = "Photo Management"

            # Add headers (starting from row 1)
            headers = ["Photo Name", "File Path", "Label", "X Coordinate", "Y Coordinate", "Direction", "Photo Time", "Visible", "Notes"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Set column widths
            column_widths = [20, 50, 20, 15, 15, 12, 20, 10, 30]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width

            # Get all photos from the photo layer (which represents imported photos)
            current_row = 2  # Data starts from row 2 (after headers)
            exported_count = 0

            for feature in photo_layer.dataProvider().getFeatures():
                feat_id = feature.id()
                photo_path = feature.attribute('path')
                photo_name = os.path.basename(photo_path) if photo_path else ""
                label_text = feature.attribute('label_text') if photo_layer.fields().indexOf('label_text') >= 0 else ""
                group_name = feature.attribute('group') if photo_layer.fields().indexOf('group') >= 0 else ""
                x_coord = feature.attribute('longitude') or ""
                y_coord = feature.attribute('latitude') or ""
                direction = feature.attribute('direction') or ""
                photo_time = feature.attribute('photo_time') if photo_layer.fields().indexOf('photo_time') >= 0 else ""

                # Determine visibility from the 'visible' attribute
                is_visible = feature.attribute('visible')
                if is_visible is None:
                    is_visible = True  # Default to visible
                visibility_text = "Yes" if is_visible else "No"

                # For exported data, use current time as import and update dates
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                ws.cell(row=current_row, column=1, value=photo_name)
                ws.cell(row=current_row, column=2, value=photo_path)
                ws.cell(row=current_row, column=3, value=label_text)
                ws.cell(row=current_row, column=4, value=x_coord)
                ws.cell(row=current_row, column=5, value=y_coord)
                ws.cell(row=current_row, column=6, value=direction)
                ws.cell(row=current_row, column=7, value=photo_time)  # Photo Time
                ws.cell(row=current_row, column=8, value=visibility_text)  # Visible
                ws.cell(row=current_row, column=9, value="")  # Notes

                current_row += 1
                exported_count += 1

            # Add data validation (dropdown) for Visibility column (column 8)
            dv_visibility = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
            dv_visibility.error = 'Please select Yes or No'
            dv_visibility.errorTitle = 'Invalid Value'
            dv_visibility.prompt = 'Select Yes or No'
            dv_visibility.promptTitle = 'Visibility'
            ws.add_data_validation(dv_visibility)
            # Apply to all rows in column 8 (Visible)
            dv_visibility.add(f'H2:H{current_row}')

            # Create Settings sheet
            ws_settings = wb.create_sheet(title="Settings")
            ws_settings.cell(row=1, column=1, value="Setting")
            ws_settings.cell(row=1, column=2, value="Value")
            ws_settings.cell(row=1, column=1).font = Font(bold=True)
            ws_settings.cell(row=1, column=2).font = Font(bold=True)
            ws_settings.column_dimensions['A'].width = 25
            ws_settings.column_dimensions['B'].width = 20
            
            # Export label settings from dialog if available
            if dlg and hasattr(dlg, 'spinBox_font_size'):
                settings_row = 2
                ws_settings.cell(row=settings_row, column=1, value="Font Size")
                ws_settings.cell(row=settings_row, column=2, value=dlg.spinBox_font_size.value())
                settings_row += 1
                
                if hasattr(dlg, 'checkBox_font_bold'):
                    ws_settings.cell(row=settings_row, column=1, value="Font Bold")
                    ws_settings.cell(row=settings_row, column=2, value="Yes" if dlg.checkBox_font_bold.isChecked() else "No")
                    font_bold_row = settings_row
                    settings_row += 1
                
                if hasattr(dlg, 'font_color'):
                    ws_settings.cell(row=settings_row, column=1, value="Font Color")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.font_color.name())
                    settings_row += 1
                
                if hasattr(dlg, 'slider_font_opacity'):
                    ws_settings.cell(row=settings_row, column=1, value="Font Opacity")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.slider_font_opacity.value())
                    settings_row += 1
                
                if hasattr(dlg, 'spinBox_buffer_size'):
                    ws_settings.cell(row=settings_row, column=1, value="Buffer Size")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.spinBox_buffer_size.value())
                    settings_row += 1
                
                if hasattr(dlg, 'buffer_color'):
                    ws_settings.cell(row=settings_row, column=1, value="Buffer Color")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.buffer_color.name())
                    settings_row += 1
                
                if hasattr(dlg, 'slider_buffer_opacity'):
                    ws_settings.cell(row=settings_row, column=1, value="Buffer Opacity")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.slider_buffer_opacity.value())
                    settings_row += 1
                
                if hasattr(dlg, 'checkBox_show_label'):
                    ws_settings.cell(row=settings_row, column=1, value="Show Label")
                    ws_settings.cell(row=settings_row, column=2, value="Yes" if dlg.checkBox_show_label.isChecked() else "No")
                    show_label_row = settings_row
                    settings_row += 1
                
                if hasattr(dlg, 'spinBox_icon_size'):
                    ws_settings.cell(row=settings_row, column=1, value="Icon Size (%)")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.spinBox_icon_size.value())
                    settings_row += 1

                # Export Show Direction preference
                if hasattr(dlg, 'checkBox_show_direction'):
                    ws_settings.cell(row=settings_row, column=1, value="Show Direction")
                    ws_settings.cell(row=settings_row, column=2, value="Yes" if dlg.checkBox_show_direction.isChecked() else "No")
                    show_direction_row = settings_row
                    settings_row += 1

                # Export icon appearance selection (camera icon file)
                if hasattr(dlg, 'comboBox_icon_appearance'):
                    try:
                        icon_choice = dlg.comboBox_icon_appearance.currentData()
                        ws_settings.cell(row=settings_row, column=1, value="Icon Appearance")
                        ws_settings.cell(row=settings_row, column=2, value=icon_choice)
                        settings_row += 1
                    except Exception:
                        # Ignore if the combobox isn't present or data not set
                        pass
                
                if hasattr(dlg, 'spinBox_tolerance'):
                    ws_settings.cell(row=settings_row, column=1, value="Click Tolerance")
                    ws_settings.cell(row=settings_row, column=2, value=dlg.spinBox_tolerance.value())
                    settings_row += 1
                
                # Add data validation (dropdown) for Font Bold in Settings sheet
                if 'font_bold_row' in locals():
                    dv_font_bold = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
                    dv_font_bold.error = 'Please select Yes or No'
                    dv_font_bold.errorTitle = 'Invalid Value'
                    dv_font_bold.prompt = 'Select Yes or No'
                    dv_font_bold.promptTitle = 'Font Bold'
                    ws_settings.add_data_validation(dv_font_bold)
                    dv_font_bold.add(f'B{font_bold_row}:B{font_bold_row}')
                
                # Add data validation (dropdown) for Show Label in Settings sheet
                if 'show_label_row' in locals():
                    dv_show_label = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
                    dv_show_label.error = 'Please select Yes or No'
                    dv_show_label.errorTitle = 'Invalid Value'
                    dv_show_label.prompt = 'Select Yes or No'
                    dv_show_label.promptTitle = 'Show Label'
                    ws_settings.add_data_validation(dv_show_label)
                    dv_show_label.add(f'B{show_label_row}:B{show_label_row}')
                # Add data validation (dropdown) for Show Direction in Settings sheet
                if 'show_direction_row' in locals():
                    dv_show_dir = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
                    dv_show_dir.error = 'Please select Yes or No'
                    dv_show_dir.errorTitle = 'Invalid Value'
                    dv_show_dir.prompt = 'Select Yes or No'
                    dv_show_dir.promptTitle = 'Show Direction'
                    ws_settings.add_data_validation(dv_show_dir)
                    dv_show_dir.add(f'B{show_direction_row}:B{show_direction_row}')
            
            # Create Advanced Labeling sheet - export full label configuration from layer
            if photo_layer and photo_layer.isValid():
                from qgis.core import QgsReadWriteContext
                from qgis.PyQt.QtXml import QDomDocument
                
                ws_advanced = wb.create_sheet(title="Advanced Labeling")
                ws_advanced.cell(row=1, column=1, value="Property")
                ws_advanced.cell(row=1, column=2, value="Value")
                ws_advanced.cell(row=1, column=1).font = Font(bold=True)
                ws_advanced.cell(row=1, column=2).font = Font(bold=True)
                ws_advanced.column_dimensions['A'].width = 30
                ws_advanced.column_dimensions['B'].width = 100
                
                # Get the layer's labeling configuration
                labeling = photo_layer.labeling()
                if labeling:
                    # Serialize the labeling configuration to XML
                    doc = QDomDocument("labeling")
                    context = QgsReadWriteContext()
                    elem = labeling.save(doc, context)
                    doc.appendChild(elem)
                    
                    # Store the XML as a string
                    xml_string = doc.toString()
                    ws_advanced.cell(row=2, column=1, value="Labeling_XML")
                    ws_advanced.cell(row=2, column=2, value=xml_string)
                    
                    # Also store whether labels are enabled
                    ws_advanced.cell(row=3, column=1, value="Labels_Enabled")
                    ws_advanced.cell(row=3, column=2, value="Yes" if photo_layer.labelsEnabled() else "No")
                    
                    QgsMessageLog.logMessage(f"[EXPORT] Advanced labeling configuration exported", 'Photo Plugin', Qgis.Info)
                else:
                    ws_advanced.cell(row=2, column=1, value="Labeling_XML")
                    ws_advanced.cell(row=2, column=2, value="None")
                    QgsMessageLog.logMessage(f"[EXPORT] No labeling configuration found on layer", 'Photo Plugin', Qgis.Warning)

            wb.save(save_path)
            wb.close()
            
            # Clean up temp file if it was created
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass  # Ignore cleanup errors

            QMessageBox.information(
                self.iface.mainWindow(),
                "Export Successful",
                f"Successfully exported {exported_count} photos to:\n{save_path}"
            )

        except Exception as e:
            # Clean up temp file on error too
            if 'temp_path' in locals() and temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            
            QMessageBox.critical(
                self.iface.mainWindow(),
                "Export Error",
                f"Failed to export Excel file: {str(e)}"
            )

    def import_plugin_state(self, layer_manager, photo_layer, photo_list_manager, symbol_renderer, label_manager=None, dlg=None, group_manager=None):
        """Import photos from an Excel file and create a new layer."""
        if not EXCEL_AVAILABLE:
            self.iface.messageBar().pushMessage(
                "Missing Dependency",
                "Excel import requires the openpyxl library. Install it with: pip install openpyxl",
                level=Qgis.Warning,
                duration=10
            )
            return

        QgsMessageLog.logMessage("[EXCEL IMPORT] Starting import process", 'Photo Plugin', Qgis.Info)
        
        # Get Excel file path from user
        file_path, _ = QFileDialog.getOpenFileName(
            self.iface.mainWindow(),
            "Import from Excel File",
            "",
            "Excel files (*.xlsx);;All files (*.*)"
        )

        if not file_path:
            QgsMessageLog.logMessage("[EXCEL IMPORT] No file selected, aborting", 'Photo Plugin', Qgis.Info)
            return

        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Selected file: {file_path}", 'Photo Plugin', Qgis.Info)

        try:
            # Check if the file is open
            if self._is_excel_file_open(file_path):
                QMessageBox.warning(
                    self.iface.mainWindow(),
                    "File In Use",
                    "The Excel file is currently open. Please close it before importing."
                )
                return

            # Read the Excel file
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", DeprecationWarning)
                    wb = load_workbook(file_path)
            except Exception as e:
                QMessageBox.critical(
                    self.iface.mainWindow(),
                    "Excel Import Error",
                    f"Failed to load Excel file. It may be open in another application or corrupted.\nError: {str(e)}"
                )
                return
            
            # Always use the "Photo Management" sheet, not the active sheet
            # This prevents issues when Settings sheet was the last active sheet
            if "Photo Management" in wb.sheetnames:
                ws = wb["Photo Management"]
            else:
                ws = wb.active
            
            QgsMessageLog.logMessage(f"[EXCEL IMPORT] Using sheet: {ws.title}", 'Photo Plugin', Qgis.Info)

            # Detect header row position
            # Check if row 1 contains "Photo Name" header (format from export_plugin_state)
            # or if it contains a title like "Visibility" (format from _create_excel_file)
            header_row = 1
            first_cell = ws.cell(row=1, column=1).value
            if first_cell and str(first_cell).strip() not in ["Photo Name", "Photo name", "photo name"]:
                # Row 1 might be a title, check row 2
                second_row_first_cell = ws.cell(row=2, column=1).value
                if second_row_first_cell and "Photo" in str(second_row_first_cell):
                    header_row = 2
            
            data_start_row = header_row + 1
            QgsMessageLog.logMessage(f"[EXCEL IMPORT] Header at row {header_row}, data starts at row {data_start_row}", 'Photo Plugin', Qgis.Info)

            # Parse the data (skip header row)
            photos_data = []
            for row in range(data_start_row, ws.max_row + 1):
                photo_name = ws.cell(row=row, column=1).value
                photo_path = ws.cell(row=row, column=2).value
                label_text = ws.cell(row=row, column=3).value  # Label column
                x_coord = ws.cell(row=row, column=4).value
                y_coord = ws.cell(row=row, column=5).value
                direction = ws.cell(row=row, column=6).value
                photo_time = ws.cell(row=row, column=7).value  # Photo Time column
                visible_text = ws.cell(row=row, column=8).value  # Visible column (column 8)

                if photo_path and x_coord is not None and y_coord is not None:
                    # Convert visible text to boolean (Yes = True/visible, No = False/hidden)
                    original_visible = True  # Default to visible
                    if visible_text:
                        visible_str = str(visible_text).strip().lower()
                        original_visible = visible_str == "yes"  # Only "yes" is visible, everything else (including "no") is hidden

                    photos_data.append({
                        'path': photo_path,
                        'label_text': str(label_text) if label_text else '',
                        'group': '',  # No longer using group
                        'latitude': float(y_coord),
                        'longitude': float(x_coord),
                        'direction': float(direction) if direction is not None else None,
                        'photo_time': str(photo_time) if photo_time else '',
                        'visible': True,  # Always import as visible
                        'original_visible': original_visible
                    })

            # Count photos with original_visible = False
            hidden_count = sum(1 for p in photos_data if not p.get('original_visible', True))
            QgsMessageLog.logMessage(f"Photos marked as hidden on import: {hidden_count}", 'Photo Plugin', Qgis.Info)
            self.iface.messageBar().pushMessage("Info", f"Imported {len(photos_data)} photos, {hidden_count} marked as hidden.", level=Qgis.Info, duration=10)

            # Read settings from Settings sheet if available
            imported_settings = {}
            if "Settings" in wb.sheetnames:
                QgsMessageLog.logMessage("\n" + "="*80, 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("[DEBUG STEP 1] Reading Settings sheet from Excel", 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("="*80, 'Photo Plugin', Qgis.Critical)
                
                ws_settings = wb["Settings"]
                for row in range(2, ws_settings.max_row + 1):
                    setting_name = ws_settings.cell(row=row, column=1).value
                    setting_value = ws_settings.cell(row=row, column=2).value
                    if setting_name and setting_value is not None:
                        imported_settings[setting_name] = setting_value
                        QgsMessageLog.logMessage(f"[DEBUG STEP 1] Read from Excel: '{setting_name}' = '{setting_value}'", 'Photo Plugin', Qgis.Critical)
                
                QgsMessageLog.logMessage(f"[DEBUG STEP 1] Complete imported_settings dict: {imported_settings}", 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage(f"[DEBUG STEP 1] 'Show Label' in dict: {'Show Label' in imported_settings}", 'Photo Plugin', Qgis.Critical)
                if "Show Label" in imported_settings:
                    QgsMessageLog.logMessage(f"[DEBUG STEP 1] 'Show Label' value: '{imported_settings['Show Label']}' (type: {type(imported_settings['Show Label'])})", 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("="*80 + "\n", 'Photo Plugin', Qgis.Critical)

            # Determine icon appearance choice from imported settings (if any)
            imported_icon_choice = None
            if "Icon Appearance" in imported_settings:
                imported_icon_choice = imported_settings.get("Icon Appearance")
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Found Icon Appearance setting: {imported_icon_choice}", 'Photo Plugin', Qgis.Info)

            if not photos_data:
                QMessageBox.warning(
                    self.iface.mainWindow(),
                    "No Data Found",
                    "No valid photo data found in the Excel file."
                )
                return

            # Ask user for confirmation with warning about clearing progress
            reply = QMessageBox.question(
                self.iface.mainWindow(),
                "Import Confirmation",
                f"⚠️ WARNING: Importing will remove all current progress and create a new layer with {len(photos_data)} photos from the Excel file.\n\nAll existing imported photos will be lost.\n\nDo you want to continue?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply != QMessageBox.Yes:
                return

            # Extract unique groups from imported data and reset group manager
            if group_manager:
                unique_groups = set()
                for photo in photos_data:
                    group = photo.get('group', '').strip()
                    if group:  # Only add non-empty groups
                        unique_groups.add(group)
                
                # Clear existing groups and add new ones from Excel
                group_manager.groups = []  # Reset groups list
                for group in sorted(unique_groups):  # Sort alphabetically
                    group_manager.add_group(group)
                
                # Save the updated groups
                group_manager._save_groups()
                group_manager.groupsChanged.emit()
                
                QgsMessageLog.logMessage(
                    f"[EXCEL IMPORT] Reset groups. Found {len(unique_groups)} unique groups: {sorted(unique_groups)}",
                    'Photo Plugin', Qgis.Info
                )

            # Always create a new photo layer for import (fresh state)
            # Pass through main plugin to get proper save path
            if self.main_plugin:
                QgsMessageLog.logMessage("[EXCEL IMPORT] Using main plugin to create layer", 'Photo Plugin', Qgis.Info)
                photo_layer = self.main_plugin.create_point_layer()
                if photo_layer:
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Layer created via main plugin: {photo_layer.name()}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Layer source: {photo_layer.source()}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Layer provider: {photo_layer.providerType()}", 'Photo Plugin', Qgis.Info)
            else:
                QgsMessageLog.logMessage("[EXCEL IMPORT] WARNING: main_plugin not available, creating memory layer", 'Photo Plugin', Qgis.Warning)
                photo_layer = layer_manager.create_point_layer(self.iface, None, None)
            if not photo_layer:
                raise ValueError("Failed to create photo layer")
            
            # DEBUG: Log layer CRS information
            layer_crs = photo_layer.crs()
            canvas_crs = self.iface.mapCanvas().mapSettings().destinationCrs()
            QgsMessageLog.logMessage(f"[DEBUG] Photo layer CRS: {layer_crs.authid()} - {layer_crs.description()}", 'Photo Plugin', Qgis.Info)
            QgsMessageLog.logMessage(f"[DEBUG] Canvas CRS: {canvas_crs.authid()} - {canvas_crs.description()}", 'Photo Plugin', Qgis.Info)

            # Initialize photo list manager if not already done
            if not photo_list_manager:
                from .photo_list_widgets import PhotoListManager
                photo_list_manager = PhotoListManager(self.iface, photo_layer)
                # Insert it into the dialog's imported photos tab
                if hasattr(self.iface.mainWindow().findChild(QWidget, "AdvancedPhotoImporterDialog"), 'imported_photos_main_layout'):
                    dlg = self.iface.mainWindow().findChild(QWidget, "AdvancedPhotoImporterDialog")
                    dlg.imported_photos_main_layout.addWidget(photo_list_manager)

            # Import each photo
            imported_count = 0
            imported_photos_data = []
            visibility_settings = []  # Store visibility settings to apply after commit
            features_to_select = []  # Features that should be selected (hidden)
            
            # DEBUG: Track coordinate ranges
            min_lat, max_lat = None, None
            min_lon, max_lon = None, None

            for photo_data in photos_data:
                try:
                    lat = photo_data.get('latitude')
                    lon = photo_data.get('longitude')
                    photo_path = photo_data.get('path')
                    direction = photo_data.get('direction')
                    group_name = photo_data.get('group', '')
                    label_text = photo_data.get('label_text', '')
                    is_visible = photo_data.get('visible', True)  # Default to visible
                    original_visible = photo_data.get('original_visible', True)
                    photo_time = photo_data.get('photo_time', '')

                    if lat is not None and lon is not None and photo_path:
                        # DEBUG: Track coordinate ranges
                        if min_lat is None or lat < min_lat: min_lat = lat
                        if max_lat is None or lat > max_lat: max_lat = lat
                        if min_lon is None or lon < min_lon: min_lon = lon
                        if max_lon is None or lon > max_lon: max_lon = lon
                        
                        # Add the point to the map, passing imported icon choice if available
                        feature = layer_manager.add_point_to_map(
                            photo_layer, lat, lon, photo_path, direction, is_visible, group_name, label_text, photo_time, svg_icon_filename=imported_icon_choice
                        )
                        if feature is not None:
                            feature_id = feature.id()
                            # Store visibility setting to apply after commit
                            visibility_settings.append((feature_id, original_visible))
                            
                            # If not visible, mark for selection and set icon to Invisible.svg
                            if not original_visible:
                                features_to_select.append(feature_id)

                            imported_count += 1
                            # Collect data for Excel
                            imported_photos_data.append({
                                'path': photo_path,
                                'x': lon,
                                'y': lat,
                                'direction': direction,
                                'photo_time': photo_time
                            })

                except Exception as e:
                    self.iface.messageBar().pushMessage(
                        "Import Warning",
                        f"Failed to import photo {photo_data.get('path', 'unknown')}: {str(e)}",
                        level=Qgis.Warning,
                        duration=5
                    )

            # DEBUG: Log coordinate ranges
            QgsMessageLog.logMessage(f"[DEBUG] Imported coordinates range:", 'Photo Plugin', Qgis.Info)
            QgsMessageLog.logMessage(f"[DEBUG]   Latitude: {min_lat} to {max_lat}", 'Photo Plugin', Qgis.Info)
            QgsMessageLog.logMessage(f"[DEBUG]   Longitude: {min_lon} to {max_lon}", 'Photo Plugin', Qgis.Info)
            QgsMessageLog.logMessage(f"[DEBUG]   Center: ({(min_lon + max_lon) / 2}, {(min_lat + max_lat) / 2})", 'Photo Plugin', Qgis.Info)

            # Commit the layer changes
            if photo_layer.isEditable():
                success = photo_layer.commitChanges()
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Layer commit: {'SUCCESS' if success else 'FAILED'}", 'Photo Plugin', Qgis.Info)
            else:
                QgsMessageLog.logMessage("[EXCEL IMPORT] Layer was not editable", 'Photo Plugin', Qgis.Warning)

            QgsMessageLog.logMessage(f"[EXCEL IMPORT] About to apply visibility to {len(visibility_settings)} features", 'Photo Plugin', Qgis.Info)
            QgsMessageLog.logMessage(f"[EXCEL IMPORT] Features to select (hide): {features_to_select}", 'Photo Plugin', Qgis.Info)

            # STEP 1: Update the layer renderer to show all photos with their SVG icons
            QgsMessageLog.logMessage("[EXCEL IMPORT] STEP 1: Setting up renderer for all photos", 'Photo Plugin', Qgis.Info)
            symbol_renderer.update_layer_symbol_manually(photo_layer, self.iface)
            self.iface.mapCanvas().refresh()
            
            QgsMessageLog.logMessage(f"[EXCEL IMPORT] All {imported_count} photos are now visible with correct SVG icons", 'Photo Plugin', Qgis.Info)
            
            # STEP 2: Apply visibility changes and select hidden photos
            QgsMessageLog.logMessage("[EXCEL IMPORT] STEP 2: Applying visibility changes for hidden photos", 'Photo Plugin', Qgis.Info)
            
            # Apply visibility attribute (for future reference)
            if visibility_settings:
                photo_layer.startEditing()
                visible_idx = photo_layer.fields().indexOf('visible')
                
                if visible_idx >= 0:
                    for feature_id, is_visible in visibility_settings:
                        photo_layer.changeAttributeValue(feature_id, visible_idx, is_visible)
                        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Set feature {feature_id} visible attribute to {is_visible}", 'Photo Plugin', Qgis.Info)
                
                photo_layer.commitChanges()
                hidden_count = sum(1 for _, is_visible in visibility_settings if not is_visible)
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Visibility attribute applied: {hidden_count} photos marked as hidden", 'Photo Plugin', Qgis.Info)
            
            # Now select features that should be hidden (this will trigger the visibility handler)
            if features_to_select:
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Selecting {len(features_to_select)} features to hide them", 'Photo Plugin', Qgis.Info)
                photo_layer.selectByIds(features_to_select)
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Selected features: {photo_layer.selectedFeatureIds()}", 'Photo Plugin', Qgis.Info)
                self.iface.messageBar().pushMessage("Info", f"Imported {imported_count} photos, {len(features_to_select)} hidden", level=Qgis.Info, duration=5)

            self.iface.mapCanvas().refresh()
            
            # Process events to ensure the layer is fully rendered before zooming
            QApplication.processEvents()

            # Zoom to the layer extent to show all imported photos
            if photo_layer and photo_layer.isValid():
                extent = photo_layer.extent()
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] ========== ZOOM DEBUG START ==========", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Layer extent: {extent.toString()}", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Extent isEmpty: {extent.isEmpty()}", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Extent center: ({extent.center().x()}, {extent.center().y()})", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Extent width: {extent.width()}, height: {extent.height()}", 'Photo Plugin', Qgis.Info)
                
                # Check current canvas state before zoom
                current_extent = self.iface.mapCanvas().extent()
                current_center = current_extent.center()
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas extent BEFORE: {current_extent.toString()}", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas center BEFORE: ({current_center.x()}, {current_center.y()})", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas scale BEFORE: 1:{int(self.iface.mapCanvas().scale())}", 'Photo Plugin', Qgis.Info)
                
                if not extent.isEmpty():
                    # Add buffer to show all photos comfortably
                    buffer_percentage = 0.3  # 30% buffer around the photos
                    extent_with_buffer = extent.buffered(extent.width() * buffer_percentage)
                    
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Buffered extent (in layer CRS): {extent_with_buffer.toString()}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Buffered center: ({extent_with_buffer.center().x()}, {extent_with_buffer.center().y()})", 'Photo Plugin', Qgis.Info)
                    
                    # Transform extent from layer CRS to canvas CRS
                    layer_crs = photo_layer.crs()
                    canvas_crs = self.iface.mapCanvas().mapSettings().destinationCrs()
                    
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Transforming extent from {layer_crs.authid()} to {canvas_crs.authid()}", 'Photo Plugin', Qgis.Info)
                    
                    if layer_crs != canvas_crs:
                        transform = QgsCoordinateTransform(layer_crs, canvas_crs, QgsProject.instance())
                        extent_transformed = transform.transformBoundingBox(extent_with_buffer)
                        QgsMessageLog.logMessage(f"[DEBUG ZOOM] Transformed extent: {extent_transformed.toString()}", 'Photo Plugin', Qgis.Info)
                        QgsMessageLog.logMessage(f"[DEBUG ZOOM] Transformed center: ({extent_transformed.center().x()}, {extent_transformed.center().y()})", 'Photo Plugin', Qgis.Info)
                    else:
                        extent_transformed = extent_with_buffer
                        QgsMessageLog.logMessage(f"[DEBUG ZOOM] No transformation needed (same CRS)", 'Photo Plugin', Qgis.Info)
                    
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Setting canvas extent...", 'Photo Plugin', Qgis.Info)
                    
                    # Set the transformed extent
                    self.iface.mapCanvas().setExtent(extent_transformed)
                    self.iface.mapCanvas().refresh()
                    QApplication.processEvents()
                    
                    # Check what happened after setting extent
                    new_extent = self.iface.mapCanvas().extent()
                    new_center = new_extent.center()
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas extent AFTER setExtent: {new_extent.toString()}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas center AFTER setExtent: ({new_center.x()}, {new_center.y()})", 'Photo Plugin', Qgis.Info)
                    
                    # Check the current scale and ensure it's not more zoomed in than 1:1200
                    # In QGIS, smaller scale values = more zoomed in
                    current_scale = self.iface.mapCanvas().scale()
                    min_scale = 1200  # Minimum scale (maximum zoom in allowed is 1:1200)
                    
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Current scale after extent: 1:{int(current_scale)}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] Min allowed scale: 1:{min_scale}", 'Photo Plugin', Qgis.Info)
                    
                    if current_scale < min_scale:
                        # Too zoomed in, zoom out to 1:1200
                        QgsMessageLog.logMessage(f"[DEBUG ZOOM] Scale too small, zooming to 1:{min_scale}", 'Photo Plugin', Qgis.Info)
                        self.iface.mapCanvas().zoomScale(min_scale)
                        self.iface.mapCanvas().refresh()
                        QApplication.processEvents()
                        
                        # Check final state after scale adjustment
                        adjusted_extent = self.iface.mapCanvas().extent()
                        adjusted_center = adjusted_extent.center()
                        QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas extent AFTER zoomScale: {adjusted_extent.toString()}", 'Photo Plugin', Qgis.Info)
                        QgsMessageLog.logMessage(f"[DEBUG ZOOM] Canvas center AFTER zoomScale: ({adjusted_center.x()}, {adjusted_center.y()})", 'Photo Plugin', Qgis.Info)
                    
                    final_scale = self.iface.mapCanvas().scale()
                    final_extent = self.iface.mapCanvas().extent()
                    final_center = final_extent.center()
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] FINAL scale: 1:{int(final_scale)}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] FINAL extent: {final_extent.toString()}", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] FINAL center: ({final_center.x()}, {final_center.y()})", 'Photo Plugin', Qgis.Info)
                    QgsMessageLog.logMessage(f"[DEBUG ZOOM] ========== ZOOM DEBUG END ==========", 'Photo Plugin', Qgis.Info)
                else:
                    QgsMessageLog.logMessage("[DEBUG ZOOM] Layer extent is empty, cannot zoom", 'Photo Plugin', Qgis.Warning)

            # Refresh the photo list if it exists
            if photo_list_manager:
                # Get the photo paths that should be hidden (before commit, feature IDs will change)
                hidden_photo_paths = []
                for photo_data in photos_data:
                    if not photo_data.get('original_visible', True):
                        hidden_photo_paths.append(photo_data['path'])
                
                QgsMessageLog.logMessage(f"[DEBUG] Hidden photo paths: {hidden_photo_paths}", 'Photo Plugin', Qgis.Info)
                QgsMessageLog.logMessage(f"[DEBUG] Currently selected features before populate_list: {photo_layer.selectedFeatureIds()}", 'Photo Plugin', Qgis.Info)
                
                photo_list_manager.setLayer(photo_layer)
                photo_list_manager.populate_list()
                
                # After populating, ensure the hidden features are still selected
                QgsMessageLog.logMessage(f"[DEBUG] Currently selected features after populate_list: {photo_layer.selectedFeatureIds()}", 'Photo Plugin', Qgis.Info)
                
                # MANUALLY uncheck the checkboxes for hidden photos by matching paths
                # This will trigger the visibility toggle signal and properly hide them
                QgsMessageLog.logMessage(f"[DEBUG] Manually unchecking checkboxes for {len(hidden_photo_paths)} hidden photos", 'Photo Plugin', Qgis.Info)
                photo_list_manager.manually_uncheck_features_by_paths(hidden_photo_paths)
                
                # Process events to ensure all signals are handled
                QApplication.processEvents()
                
                QgsMessageLog.logMessage("[EXCEL IMPORT] Photo list manager updated", 'Photo Plugin', Qgis.Info)

            # Update Excel file with imported photos
            if imported_photos_data:
                self._update_excel_file(imported_photos_data)

            # Final verification of hidden photos
            final_selected = photo_layer.selectedFeatureIds()
            QgsMessageLog.logMessage(f"[DEBUG] FINAL CHECK - Selected (hidden) features: {final_selected}", 'Photo Plugin', Qgis.Info)
            QgsMessageLog.logMessage(f"[DEBUG] FINAL CHECK - Total features: {photo_layer.featureCount()}", 'Photo Plugin', Qgis.Info)
            
            # Verify each selected feature has visible=False
            for feature in photo_layer.getFeatures():
                feat_id = feature.id()
                if feat_id in final_selected:
                    visible_val = feature.attribute('visible')
                    QgsMessageLog.logMessage(f"[DEBUG] Feature {feat_id} visible attribute: {visible_val}", 'Photo Plugin', Qgis.Info)

            # Apply imported settings to dialog and label manager
            QgsMessageLog.logMessage("\n" + "="*80, 'Photo Plugin', Qgis.Critical)
            QgsMessageLog.logMessage("[DEBUG STEP 2] Starting to apply settings to dialog", 'Photo Plugin', Qgis.Critical)
            QgsMessageLog.logMessage(f"[DEBUG STEP 2] imported_settings exists: {bool(imported_settings)}", 'Photo Plugin', Qgis.Critical)
            QgsMessageLog.logMessage(f"[DEBUG STEP 2] dlg exists: {bool(dlg)}", 'Photo Plugin', Qgis.Critical)
            QgsMessageLog.logMessage(f"[DEBUG STEP 2] label_manager exists: {bool(label_manager)}", 'Photo Plugin', Qgis.Critical)
            QgsMessageLog.logMessage("="*80, 'Photo Plugin', Qgis.Critical)
            
            if imported_settings and dlg and label_manager:
                from qgis.PyQt.QtGui import QColor
                
                QgsMessageLog.logMessage(f"[EXCEL IMPORT LABELS] Applying imported settings to dialog and label manager", 'Photo Plugin', Qgis.Info)
                
                # Apply to dialog
                if "Font Size" in imported_settings and hasattr(dlg, 'spinBox_font_size'):
                    dlg.spinBox_font_size.setValue(float(imported_settings["Font Size"]))
                
                if "Font Bold" in imported_settings and hasattr(dlg, 'checkBox_font_bold'):
                    dlg.checkBox_font_bold.setChecked(imported_settings["Font Bold"] == "Yes")
                
                if "Font Color" in imported_settings and hasattr(dlg, 'font_color'):
                    color = QColor(imported_settings["Font Color"])
                    dlg.font_color = color
                    dlg.pushButton_font_color.setStyleSheet(f"background-color: {color.name()};")
                
                if "Font Opacity" in imported_settings and hasattr(dlg, 'slider_font_opacity'):
                    dlg.slider_font_opacity.setValue(int(imported_settings["Font Opacity"]))
                
                if "Buffer Size" in imported_settings and hasattr(dlg, 'spinBox_buffer_size'):
                    dlg.spinBox_buffer_size.setValue(float(imported_settings["Buffer Size"]))
                
                if "Buffer Color" in imported_settings and hasattr(dlg, 'buffer_color'):
                    color = QColor(imported_settings["Buffer Color"])
                    dlg.buffer_color = color
                    dlg.pushButton_buffer_color.setStyleSheet(f"background-color: {color.name()};")
                
                if "Buffer Opacity" in imported_settings and hasattr(dlg, 'slider_buffer_opacity'):
                    dlg.slider_buffer_opacity.setValue(int(imported_settings["Buffer Opacity"]))
                
                if "Click Tolerance" in imported_settings and hasattr(dlg, 'spinBox_tolerance'):
                    tolerance_value = float(imported_settings["Click Tolerance"])
                    dlg.spinBox_tolerance.setValue(int(tolerance_value))
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Click Tolerance setting from Excel: {tolerance_value}", 'Photo Plugin', Qgis.Info)
                
                if ("Icon Size" in imported_settings or "Icon Size (%)" in imported_settings) and hasattr(dlg, 'spinBox_icon_size'):
                    icon_size_value = int(imported_settings.get("Icon Size (%)", imported_settings.get("Icon Size", 100)))
                    dlg.spinBox_icon_size.setValue(icon_size_value)
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Icon Size setting from Excel: {icon_size_value}%", 'Photo Plugin', Qgis.Info)

                # Apply Show Direction setting from Excel (if present)
                if "Show Direction" in imported_settings and hasattr(dlg, 'checkBox_show_direction'):
                    raw = imported_settings.get("Show Direction")
                    checked = True if str(raw).strip().lower() == 'yes' else False
                    dlg.checkBox_show_direction.setChecked(checked)
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Show Direction set to: {checked}", 'Photo Plugin', Qgis.Info)
                    # Apply to renderer via main plugin if available
                    try:
                        if hasattr(self, 'main_plugin') and self.main_plugin:
                            self.main_plugin.symbol_renderer.set_include_direction(checked)
                            # Update the renderer to apply the change
                            if photo_layer and photo_layer.isValid():
                                self.main_plugin.symbol_renderer.update_layer_symbol_manually(photo_layer, self.iface)
                    except Exception as e:
                        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Failed to apply Show Direction: {e}", 'Photo Plugin', Qgis.Warning)

                # Apply Icon Appearance setting: set dialog combobox and call main plugin to apply
                if "Icon Appearance" in imported_settings:
                    icon_choice = imported_settings.get("Icon Appearance")
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Applying Icon Appearance: {icon_choice}", 'Photo Plugin', Qgis.Info)
                    try:
                        if hasattr(dlg, 'comboBox_icon_appearance') and icon_choice:
                            combo = dlg.comboBox_icon_appearance
                            # Try to find matching item by data()
                            matched_index = -1
                            for i in range(combo.count()):
                                try:
                                    if combo.itemData(i) == icon_choice:
                                        matched_index = i
                                        break
                                except Exception:
                                    continue
                            if matched_index >= 0:
                                combo.setCurrentIndex(matched_index)
                    except Exception:
                        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Failed to set comboBox_icon_appearance to {icon_choice}", 'Photo Plugin', Qgis.Warning)

                    # Call main plugin to apply the icon appearance to the layer (updates renderer and backups)
                    try:
                        if hasattr(self, 'main_plugin') and self.main_plugin:
                            self.main_plugin.update_icon_appearance(icon_choice)
                        else:
                            # Fallback: attempt to update renderer directly
                            try:
                                symbol_renderer.update_layer_symbol_manually(photo_layer, self.iface)
                            except Exception:
                                pass
                    except Exception as e:
                        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Error applying Icon Appearance: {e}", 'Photo Plugin', Qgis.Warning)
                
                QgsMessageLog.logMessage("\n" + "*"*80, 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("[DEBUG STEP 3] Processing Show Label setting", 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("*"*80, 'Photo Plugin', Qgis.Critical)
                
                show_label = False  # Default to False
                if "Show Label" in imported_settings:
                    raw_value = imported_settings["Show Label"]
                    show_label = raw_value == "Yes"
                    QgsMessageLog.logMessage(f"[DEBUG STEP 3] Found 'Show Label' in imported_settings", 'Photo Plugin', Qgis.Critical)
                    QgsMessageLog.logMessage(f"[DEBUG STEP 3] Raw value from Excel: '{raw_value}' (type: {type(raw_value)})", 'Photo Plugin', Qgis.Critical)
                    QgsMessageLog.logMessage(f"[DEBUG STEP 3] Comparison result (raw_value == 'Yes'): {show_label}", 'Photo Plugin', Qgis.Critical)
                    QgsMessageLog.logMessage(f"[DEBUG STEP 3] Final show_label boolean: {show_label}", 'Photo Plugin', Qgis.Critical)
                    
                    if hasattr(dlg, 'checkBox_show_label'):
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] Dialog has checkBox_show_label", 'Photo Plugin', Qgis.Critical)
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] Checkbox state BEFORE setting: {dlg.checkBox_show_label.isChecked()}", 'Photo Plugin', Qgis.Critical)
                        
                        # Block signals to prevent triggering events during import
                        dlg.checkBox_show_label.blockSignals(True)
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] Signals blocked", 'Photo Plugin', Qgis.Critical)
                        
                        dlg.checkBox_show_label.setChecked(show_label)
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] Called setChecked({show_label})", 'Photo Plugin', Qgis.Critical)
                        
                        dlg.checkBox_show_label.blockSignals(False)
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] Signals unblocked", 'Photo Plugin', Qgis.Critical)
                        
                        actual_state = dlg.checkBox_show_label.isChecked()
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] Checkbox state AFTER setting: {actual_state}", 'Photo Plugin', Qgis.Critical)
                        
                        if actual_state != show_label:
                            QgsMessageLog.logMessage(f"[DEBUG STEP 3] *** ERROR *** Checkbox state doesn't match! Expected: {show_label}, Got: {actual_state}", 'Photo Plugin', Qgis.Critical)
                        else:
                            QgsMessageLog.logMessage(f"[DEBUG STEP 3] ✓ Checkbox state matches expected value", 'Photo Plugin', Qgis.Critical)
                    else:
                        QgsMessageLog.logMessage(f"[DEBUG STEP 3] *** ERROR *** Dialog does NOT have checkBox_show_label!", 'Photo Plugin', Qgis.Critical)
                else:
                    QgsMessageLog.logMessage(f"[DEBUG STEP 3] 'Show Label' NOT found in imported_settings, using default: {show_label}", 'Photo Plugin', Qgis.Critical)
                
                QgsMessageLog.logMessage("*"*80 + "\n", 'Photo Plugin', Qgis.Critical)
                
                # Update label manager settings
                font_size = float(imported_settings.get("Font Size", 7.0))
                font_bold = imported_settings.get("Font Bold", "Yes") == "Yes"
                font_color = QColor(imported_settings.get("Font Color", "#000000"))
                font_opacity = int(imported_settings.get("Font Opacity", 100)) / 100.0
                font_color.setAlphaF(font_opacity)
                
                buffer_size = float(imported_settings.get("Buffer Size", 1.0))
                buffer_color = QColor(imported_settings.get("Buffer Color", "#ffffff"))
                buffer_opacity = int(imported_settings.get("Buffer Opacity", 100)) / 100.0
                buffer_color.setAlphaF(buffer_opacity)
                
                label_manager.update_label_style(
                    font_size=font_size,
                    font_bold=font_bold,
                    font_color=font_color,
                    buffer_size=buffer_size,
                    buffer_color=buffer_color
                )
                
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Label manager settings updated (not applied yet)", 'Photo Plugin', Qgis.Info)
                
                # Commit all changes before applying settings
                if photo_layer.isEditable():
                    photo_layer.commitChanges()
            
            # Extract click tolerance if available
            click_tolerance = None
            if "Click Tolerance" in imported_settings:
                click_tolerance = float(imported_settings["Click Tolerance"])
                QgsMessageLog.logMessage(f"[EXCEL IMPORT] Extracted click tolerance: {click_tolerance}", 'Photo Plugin', Qgis.Info)
            
            # APPLY ALL SETTINGS using main plugin method
            if self.main_plugin and hasattr(self.main_plugin, 'apply_all_settings'):
                QgsMessageLog.logMessage("\n" + "="*80, 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("[EXCEL IMPORT] Calling main plugin's apply_all_settings()", 'Photo Plugin', Qgis.Critical)
                if dlg and hasattr(dlg, 'checkBox_show_label'):
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Dialog checkbox state BEFORE apply_all_settings: {dlg.checkBox_show_label.isChecked()}", 'Photo Plugin', Qgis.Critical)
                QgsMessageLog.logMessage("="*80, 'Photo Plugin', Qgis.Critical)
                
                try:
                    # Pass the new photo_layer to apply_all_settings
                    self.main_plugin.apply_all_settings(photo_layer=photo_layer)
                    
                    QgsMessageLog.logMessage("\n" + "="*80, 'Photo Plugin', Qgis.Critical)
                    QgsMessageLog.logMessage("[EXCEL IMPORT] apply_all_settings() completed", 'Photo Plugin', Qgis.Critical)
                    if photo_layer and photo_layer.isValid():
                        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Layer labelsEnabled: {photo_layer.labelsEnabled()}", 'Photo Plugin', Qgis.Critical)
                    if dlg and hasattr(dlg, 'checkBox_show_label'):
                        QgsMessageLog.logMessage(f"[EXCEL IMPORT] Dialog checkbox state AFTER apply_all_settings: {dlg.checkBox_show_label.isChecked()}", 'Photo Plugin', Qgis.Critical)
                    QgsMessageLog.logMessage("="*80 + "\n", 'Photo Plugin', Qgis.Critical)
                except Exception as e:
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] ERROR in apply_all_settings: {str(e)}", 'Photo Plugin', Qgis.Critical)
                    import traceback
                    QgsMessageLog.logMessage(f"[EXCEL IMPORT] Traceback: {traceback.format_exc()}", 'Photo Plugin', Qgis.Critical)
                    raise  # Re-raise so it's caught by the outer try-except
            else:
                QgsMessageLog.logMessage("[EXCEL IMPORT] ERROR: Cannot call apply_all_settings - main_plugin not set!", 'Photo Plugin', Qgis.Critical)

            QMessageBox.information(
                self.iface.mainWindow(),
                "Import Successful",
                f"Successfully imported {imported_count} photos from {os.path.basename(file_path)}"
            )
            
            # Return the new photo layer, photo_list_manager, and click_tolerance
            return photo_layer, photo_list_manager, click_tolerance

        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            QgsMessageLog.logMessage(f"[EXCEL IMPORT ERROR] {error_trace}", 'Photo Plugin', Qgis.Critical)
            QMessageBox.critical(
                self.iface.mainWindow(),
                "Import Error",
                f"Failed to import from Excel file: {str(e)}\n\nCheck the message log for details."
            )
            return None, None, None

    def _update_excel_file(self, photo_data_list, photo_layer=None):
        """Update the Excel file with new photo data."""
        if not EXCEL_AVAILABLE:
            return

        excel_path = self._get_excel_file_path()
        
        # Check if file is in use (locked)
        if os.path.exists(excel_path):
            try:
                # Try to open file in exclusive mode to check if it's locked
                with open(excel_path, 'r+b') as f:
                    pass
            except (IOError, PermissionError):
                QgsMessageLog.logMessage(
                    "Excel file is currently in use. Please close it and try again.",
                    'Photo Plugin',
                    Qgis.Warning
                )
                return

        # Create file if it doesn't exist
        if not os.path.exists(excel_path):
            self._create_excel_file(excel_path)
            if not os.path.exists(excel_path):
                # Creation failed
                return

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", DeprecationWarning)
                wb = load_workbook(excel_path)
            ws = wb.active

            # Create a mapping of photo paths to row numbers (data starts from row 2)
            path_to_row = {}
            for row in range(2, ws.max_row + 1):
                path_cell = ws.cell(row=row, column=2).value
                if path_cell:
                    path_to_row[path_cell] = row

            # Process photos (add new or update existing)
            current_row = ws.max_row + 1
            added_count = 0
            updated_count = 0

            for photo_data in photo_data_list:
                photo_path = photo_data.get('path', '')
                photo_name = os.path.basename(photo_path)
                x_coord = photo_data.get('x', '')
                y_coord = photo_data.get('y', '')
                direction = photo_data.get('direction', '')
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Determine visibility: find the feature and check if it's selected
                is_visible = True  # Default to visible
                if photo_layer and photo_layer.isValid():
                    for feature in photo_layer.getFeatures():
                        if feature.attribute('path') == photo_path:
                            is_visible = feature.attribute('visible')
                            if is_visible is None:
                                is_visible = True
                            break

                visibility_text = "Yes" if is_visible else "No"

                if photo_path in path_to_row:
                    # Update existing photo
                    row = path_to_row[photo_path]
                    ws.cell(row=row, column=3, value=x_coord)
                    ws.cell(row=row, column=4, value=y_coord)
                    ws.cell(row=row, column=5, value=direction)
                    ws.cell(row=row, column=7, value=current_time)  # Update Date
                    ws.cell(row=row, column=8, value=visibility_text)  # Visible column
                    updated_count += 1
                else:
                    # Add new photo
                    ws.cell(row=current_row, column=1, value=photo_name)
                    ws.cell(row=current_row, column=2, value=photo_path)
                    ws.cell(row=current_row, column=3, value=x_coord)
                    ws.cell(row=current_row, column=4, value=y_coord)
                    ws.cell(row=current_row, column=5, value=direction)
                    ws.cell(row=current_row, column=6, value=current_time)  # Import Date
                    ws.cell(row=current_row, column=7, value=current_time)  # Last Update (same as import for new photos)
                    ws.cell(row=current_row, column=8, value=visibility_text)  # Visible column
                    ws.cell(row=current_row, column=9, value="")  # Notes

                    current_row += 1
                    added_count += 1

            if added_count > 0 or updated_count > 0:
                try:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore", DeprecationWarning)
                        wb.save(excel_path)
                except Exception as save_error:
                    # Save failed - log and skip message
                    QgsMessageLog.logMessage(
                        f"Failed to save Excel file: {str(save_error)}",
                        'Photo Plugin',
                        Qgis.Critical
                    )
                    self.iface.messageBar().pushMessage(
                        "Excel Save Failed",
                        "Could not save Excel file. It may be open in another program.",
                        level=Qgis.Warning,
                        duration=5
                    )
                    return
                
                message = f"Excel file updated: {os.path.basename(excel_path)}"
                if added_count > 0:
                    message += f" (added {added_count})"
                if updated_count > 0:
                    message += f" (updated {updated_count})"
                self.iface.messageBar().pushMessage(
                    "Excel Updated",
                    message,
                    level=Qgis.Info,
                    duration=5
                )

        except Exception as e:
            self.iface.messageBar().pushMessage(
                "Excel Update Error",
                f"Failed to update Excel file: {str(e)}",
                level=Qgis.Warning,
                duration=5
            )
